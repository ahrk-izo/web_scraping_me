'''
Webスクレイピング
とあるデータを抽出し、CSVおよびExcelに出力する
'''
import re
import time
import pprint
import configparser
import requests
import csv
import datetime
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.chart import Reference
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd


def get_login_data():
    'Configファイルの読み込み、ログインに関するデータを取得する'
    login_data = {}
    env_cfg = configparser.ConfigParser()
    env_cfg.read('env.cfg')
    login_data['url'] = env_cfg['url']['login']
    login_data['contents_url'] = env_cfg['url']['contents']
    login_data['user_name'] = env_cfg['account']['username']
    login_data['pass'] = env_cfg['account']['pass']
    login_data['elem_login_id'] = env_cfg['element']['login_id']
    login_data['elem_pass_id'] = env_cfg['element']['pass_id']
    login_data['elem_btn_class'] = env_cfg['element']['btn_class']

    return login_data


def output_csv_file(dates, titles, article_length_list):
    'Output CSV File'
    output_file = open('output.csv', 'w', newline='', encoding='shift_jis')
    output_writer = csv.writer(output_file)
    output_writer.writerow(['date', 'title', 'article_length'])
    for i, date in enumerate(dates):
        output_writer.writerow([date, titles[i], article_length_list[i]])
    output_file.close()


def output_excel_file(dates, titles, article_length_list):
    'Output Excel File'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'diary_data'
    border = Border(top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'),
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'))

    # Title
    ws['A1'] = 'ブログ記事のまとめ'
    ws['A1'].font = Font(name='メイリオ', size=18, color='0000ff', bold=True)
    ws['C1'] = datetime.datetime.today().strftime('%Y/%m/%d')
    ws['C1'].alignment = Alignment(horizontal='right')

    # 項目名
    item_row = 2
    for i, item in enumerate(['date', 'title', 'article_length'], 1):  # indexは1オリジン
        cell_coordinate = ws.cell(row=item_row, column=i).coordinate
        ws[cell_coordinate].value = item
        ws[cell_coordinate].border = border

    # 各データ
    data_start_row = item_row + 1
    for i, date in enumerate(dates):
        cell_coordinate = ws.cell(row=data_start_row+i, column=1).coordinate
        ws[cell_coordinate].value = date
        ws[cell_coordinate].border = border
        cell_coordinate = ws.cell(row=data_start_row+i, column=2).coordinate
        ws[cell_coordinate].value = titles[i]
        ws[cell_coordinate].border = border
        cell_coordinate = ws.cell(row=data_start_row+i, column=3).coordinate
        ws[cell_coordinate].value = article_length_list[i]
        ws[cell_coordinate].border = border

    # 幅指定
    # cell_column = ws.cell(row=1, column=2).column
    # ws.column_dimensions[cell_column].width = 50  # B列のこと(これでもよい)
    ws.column_dimensions['B'].width = 50  # B列
    ws.column_dimensions['C'].width = 15  # C列

    # グラフ
    ref_obj = openpyxl.chart.Reference(ws, min_row=data_start_row, min_col=3, max_row=len(dates)+1, max_col=3)
    series_obj = openpyxl.chart.Series(ref_obj, title='article_length')
    chart_obj = openpyxl.chart.BarChart()
    chart_obj.style = 11  # スタイル(なんかかっこいい)
    chart_obj.type = 'bar'  # 横軸
    chart_obj.width = 18  # サイズ # default is 15
    chart_obj.height = 15  # サイズ # default is 7.5
    chart_obj.append(series_obj)
    dates = Reference(ws, min_row=data_start_row, min_col=1, max_row=len(dates)+1)  # 軸に使う範囲指定
    chart_obj.set_categories(dates)  # 軸の範囲に指定
    ws.add_chart(chart_obj, 'E2')  # 表示位置指定

    wb.save('output.xlsx')  # 保存


def web_scraping(login_data):
    'Seleniumでサイトにアクセスし、データを抽出する'

    print('Webスクレイピング開始---')
    driver = webdriver.Chrome()
    try:
        driver.get(login_data['url'])
        WebDriverWait(driver, 30).until(
            EC.presence_of_all_elements_located((By.ID, "top"))
        )
    except:
        driver.quit()
        return

    # ログイン用CSSスタイル要素取得
    elem_user_name = driver.find_element_by_id(login_data['elem_login_id'])
    elem_pass = driver.find_element_by_id(login_data['elem_pass_id'])
    elem_login_btn = driver.find_element_by_class_name(
                                    login_data['elem_btn_class'])
    time.sleep(2)
    # フォームに入力、ボタンクリック
    elem_user_name.send_keys(login_data['user_name'])
    elem_pass.send_keys(login_data['pass'])
    time.sleep(2)
    elem_login_btn.click()
    time.sleep(5)

    # コンテンツページアクセス
    try:
        driver.get(login_data['contents_url'])
        WebDriverWait(driver, 30).until(
            EC.presence_of_all_elements_located
        )
    except:
        driver.quit()
        return

    dates = []
    titles = []
    article_length_list = []
    for _ in range(5):
        # タイトル取得
        elems = driver.find_elements_by_class_name('header-title')
        _titles = [elem.text for elem in elems]
        titles.extend(_titles)

        # 日付取得
        elems = driver.find_elements_by_class_name('metadata-publish-date')
        _dates = [elem.text for elem in elems]
        dates.extend(_dates)

        # 記事
        elems = driver.find_elements_by_class_name('asset-content')
        elems = [elem.find_element_by_class_name('journal-content-article') for elem in elems]
        articles = [elem.text for elem in elems]
        # pprint.pprint(articles)
        _article_length_list = [len(article) for article in articles]
        article_length_list.extend(_article_length_list)

        # 画像
        # TODO: 画像の取得ができない
        # img_urls = [elem.find_element_by_tag_name('img').get_attribute('src') for elem in elems]
        # pprint.pprint(img_urls)

        # ページ遷移ボタン
        elem = driver.find_element_by_class_name('lfr-pagination-buttons')
        link_next_page = elem.find_element_by_link_text('次へ')  # 「次へ」のリンクを取得
        link_next_page.click()  # リンクに移動

    # pprint.pprint(dates)
    # pprint.pprint(titles)
    # pprint.pprint(article_length_list)
    time.sleep(5)

    # CSV書き出し
    print('CSVに書き出し---')
    output_csv_file(dates, titles, article_length_list)

    # Excel書き出し
    print('Excelに書き出し---')
    output_excel_file(dates, titles, article_length_list)

    driver.quit()


if __name__ == '__main__':
    'main'
    LOGIN_DATA = get_login_data()
    # pprint.pprint(LOGIN_DATA)
    web_scraping(LOGIN_DATA)

    '''
    # Excel書き出し
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'diary_data'

    ws['A1'] = 'ブログ記事のまとめ'
    ws['C1'] = datetime.datetime.today().strftime('%Y/%m/%d')
    ws['A1'].font = Font(name='メイリオ', size=18, color='0000ff', bold=True)
    ws['C1'].alignment = Alignment(horizontal='right')

    for i, item in enumerate(['date', 'title', 'article_length'], 1):  # indexは1オリジン
        cell_coordinate = ws.cell(row=2, column=i).coordinate
        ws[cell_coordinate].value = item

    wb.save('output.xlsx')
    '''
